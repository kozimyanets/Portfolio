import datetime
import openpyxl
import wx

wb=openpyxl.load_workbook('Книга1.xlsx') # загрузка тестовых данных

def showAnswer(evt):
	'''Функция для показа ответаов на вопросы ликбеза'''
	question = infoList.GetFirstSelected()
	question = infoList.GetItemText(question, 0)
	wx.MessageBox(infoDict[question], question)

def listItemReview(event, table):
	'''Функция для построения окна с подробностями предложений и тем'''
	dialog=wx.Dialog(None, -1, 'Подробности', style=wx.CAPTION, size=(750, 750))
	dialogSizer=wx.BoxSizer(wx.VERTICAL)
	dialog.SetSizer(dialogSizer)
	for i in wb[table].iter_rows(max_row=1, values_only=True): heading=i
	if table == 'Предложения': mylist = RPList
	elif table == 'Темник': mylist = temnikList
	number = mylist.GetFirstSelected()
	number = int(mylist.GetItemText(number, 0)) + 1
	for i in wb[table].iter_rows(min_row=number, max_row=number, values_only=True): values = i
	for (i, j) in zip(heading, values):
		sizer = wx.BoxSizer(wx.HORIZONTAL)
		label = wx.StaticText(dialog, -1, i+': ')
		sizer.Add(label, proportion=1, flag=wx.EXPAND)
		edit = wx.TextCtrl(dialog, -1, style=wx.TE_MULTILINE|wx.TE_READONLY)
		edit.SetValue(str(j))
		sizer.Add(edit, proportion=20, flag=wx.EXPAND)
		dialogSizer.Add(sizer, wx.EXPAND)
	OKButton = wx.Button(dialog, -1, 'OK')
	dialog.Bind(wx.EVT_BUTTON, (lambda evt: dialog.Destroy()), OKButton)
	sizer = wx.BoxSizer(wx.HORIZONTAL)
	sizer.Add(wx.Size(50, 50), proportion=6)
	sizer.Add(OKButton, proportion=1, flag=wx.EXPAND)
	if table == 'Предложения':
		changeButton = wx.Button(dialog, -1, 'Дополнить')
		dialog.Bind(wx.EVT_BUTTON, (lambda event: changeRPDialog()), changeButton)
		sizer.Add(changeButton, proportion=1, flag=wx.EXPAND)
	elif table == 'Темник':
		addButton=wx.Button(dialog, -1, 'Предложить решение')
		dialog.Bind(wx.EVT_BUTTON, (lambda event: addRPDialog(temnik=number-1)), addButton)
		sizer.Add(addButton, proportion=1, flag=wx.EXPAND)
	dialogSizer.Add(sizer, flag=wx.EXPAND)
	dialog.Layout()
	if dialog.ShowModal() == wx.ID_OK: dialog.Destroy()
	dialog.Destroy()

def updateRPList():
	'''Функция для обновления элементов списка'''
	RPList.DeleteAllItems()
	for i in wb['Предложения'].iter_rows(values_only=True):
		if str(i[2]) == person[0]: RPList.Append((i[0], i[4], i[5]))

def changeRPDialog():
	'''Функция создания и работы окна дополнения предложения'''
	number=RPList.GetFirstSelected()
	number=int(RPList.GetItemText(number, 0)) + 1
	value = wb['Предложения'].cell(row=number, column=5).value
	dialog = wx.Dialog(None, -1, title='Дополнение предложения')
	dialogSizer = wx.GridBagSizer()
	dialog.SetSizer(dialogSizer)
	label = wx.StaticText(dialog, -1, 'Предложение: ')
	dialogSizer.Add(label, pos=(0, 0), flag=wx.EXPAND)
	edit = wx.TextCtrl(dialog, -1, style=wx.TE_MULTILINE|wx.TE_READONLY)
	edit.SetValue(value)
	dialogSizer.Add(edit, pos=(0, 1), span=(2, 2), flag=wx.EXPAND)
	label=wx.StaticText(dialog, -1, 'Дополнение: ')
	dialogSizer.Add(label, pos=(2, 0), flag=wx.EXPAND)
	edit=wx.TextCtrl(dialog, -1, style=wx.TE_MULTILINE)
	dialogSizer.Add(edit, pos=(2, 1), span=(2, 2), flag=wx.EXPAND)
	edit.SetFocus()
	dialogSizer.Add(dialog.CreateButtonSizer(flags=wx.OK|wx.CANCEL), pos=(4, 2), flag=wx.EXPAND)
	if dialog.ShowModal() == wx.ID_OK:
		new = edit.GetValue()
		if new:
			value = value + '\nдополнение от ' + str(datetime.datetime.now()).split(' ')[0] + ': ' + new
			wb['Предложения'].cell(row=number, column=5).value = value
			wb.save('Книга1.xlsx')
			updateRPList()
	dialog.Destroy()

def addRPDialog(temnik='Нет'):
	'''Функция создания и работы диалога "Заявка на предложение"'''
	dialog = wx.Dialog(None, -1, title='Заявка на предложение')
	dialogSizer = wx.GridBagSizer()
	dialog.SetSizer(dialogSizer)
	label = wx.StaticText(dialog, -1, 'Предложение: ')
	dialogSizer.Add(label, pos=(0, 0), flag=wx.EXPAND)
	edit = wx.TextCtrl(dialog, -1, style=wx.TE_MULTILINE)
	dialogSizer.Add(edit, pos=(0, 1), span=(2, 2), flag=wx.EXPAND)
	edit.SetFocus()
	dialogSizer.Add(dialog.CreateButtonSizer(flags=wx.OK|wx.CANCEL), pos=(2, 2), flag=wx.EXPAND)
	if dialog.ShowModal() == wx.ID_OK:
		record = []
		m = max([i.value for i in wb['Предложения']['a']][1:])
		record.append(m+1)
		record.append(datetime.datetime.now())
		record += person
		record.append(edit.GetValue())
		record.append('Рассматривается в ОВРП')
		record.append(temnik)
		wb['Предложения'].append(record)
		wb.save('Книга1.xlsx')
		updateRPList()
	dialog.Destroy()

def createStartDialog(wb):
	'''Создает диалог входа'''
	dialog = wx.Dialog(None, -1, title='Вход', size=(300, 300))
	dialogSizer = wx.GridBagSizer()
	dialog.SetSizer(dialogSizer)
	heading1 = wx.StaticText(dialog, -1, 'ФИО:')
	dialogSizer.Add(heading1, pos=(0, 0))
	value1 = wx.TextCtrl(dialog, -1)
	dialogSizer.Add(value1, pos=(0, 1), span=(1, 2), flag=wx.EXPAND)
	heading2 = wx.StaticText(dialog, -1, 'Таб. №:')
	dialogSizer.Add(heading2, pos=(1, 0))
	value2 = wx.TextCtrl(dialog, -1)
	dialogSizer.Add(value2, pos=(1, 1), span=(1, 2), flag=wx.EXPAND)
	buttonSizer = dialog.CreateButtonSizer(flags=wx.OK|wx.CANCEL)
	dialogSizer.Add(buttonSizer, pos=(2, 0), span=(1, 3), flag=wx.EXPAND)
	answer = dialog.ShowModal()
	if answer == wx.ID_OK:
		name = value1.GetValue()
		id = value2.GetValue()
		for i in wb['Персонал'].rows:
			if str(i[0].value)==id and name==i[1].value:
				dialog.Destroy()
				return id, name
		wx.MessageBox(message='Неверное ФИО или табельный номер! Проверьте введенные данные!', caption='Внимание!')
		return False
	else:
		dialog.Destroy()
		app.__exit__()

app = wx.App()
person = False
while person == False: person=createStartDialog(wb)
mainFrame = wx.Frame(None, title='Рационализация', size=(750, 750))
notebook = wx.Notebook(mainFrame, style=wx.NB_BOTTOM)
RPPage = wx.Panel(notebook, -1)
RPSizer = wx.GridBagSizer()
RPPage.SetSizer(RPSizer)
RPList = wx.ListCtrl(RPPage, -1, style=wx.LC_REPORT|wx.LC_SINGLE_SEL)
for i in [('№', 40), ('Предмет', 500), ('Статус', 200)]: RPList.AppendColumn(i[0], width=i[1])
updateRPList()
RPList.Bind(wx.EVT_LIST_ITEM_ACTIVATED, (lambda event: listItemReview(event, 'Предложения')))
RPSizer.Add(RPList, pos=(0, 0), span=(1, 4), flag=wx.EXPAND)
changeButton = wx.Button(RPPage, -1, 'Дополнить')
RPPage.Bind(wx.EVT_BUTTON, (lambda event: changeRPDialog()), changeButton)
RPSizer.Add(changeButton, pos=(1, 3), flag=wx.EXPAND)
addButton = wx.Button(RPPage, -1, 'Новое')
RPPage.Bind(wx.EVT_BUTTON, (lambda event: addRPDialog()), addButton)
RPSizer.Add(addButton, pos=(1, 2), flag=wx.EXPAND)
infoButton=wx.Button(RPPage, -1, 'Подробнее')
RPPage.Bind(wx.EVT_BUTTON, (lambda event: listItemReview(event, 'Предложения')), infoButton)
RPSizer.Add(infoButton, pos=(1, 4), flag=wx.EXPAND)
notebook.AddPage(RPPage, 'Предложения')
temnik = wx.Panel(notebook, -1)
temnikSizer=wx.BoxSizer(wx.VERTICAL)
temnik.SetSizer(temnikSizer)
temnikList = wx.ListCtrl(temnik, -1, style=wx.LC_REPORT|wx.LC_SINGLE_SEL)
for i in [('№', 40), ('Задача', 700)]: temnikList.AppendColumn(i[0], width=i[1])
for i in wb['Темник'].iter_rows(min_row=2, values_only=True): temnikList.Append((i[0], i[2]))
temnikList.Bind(wx.EVT_LIST_ITEM_ACTIVATED, (lambda event: listItemReview(event, 'Темник')))
temnikSizer.Add(temnikList, flag=wx.EXPAND)
notebook.AddPage(temnik, 'Темник')
info = wx.Panel(notebook, -1)
infoSizer=wx.BoxSizer(wx.VERTICAL)
info.SetSizer(infoSizer)
infoList = wx.ListCtrl(info, -1, style=wx.LC_REPORT|wx.LC_NO_HEADER|wx.LC_SINGLE_SEL)
infoList.AppendColumn('Вопрос', width=740)
infoDict={}
for i in wb['Ликбез'].iter_rows(values_only=True): infoDict[i[0]]=i[1]
for i in infoDict.keys(): infoList.Append((i, ))
infoList.Bind(wx.EVT_LIST_ITEM_ACTIVATED, (lambda evt: showAnswer(evt)))
infoSizer.Add(infoList, flag=wx.EXPAND)
notebook.AddPage(info, 'Ликбез')
contacts = wx.Panel(notebook, -1)
contactsSizer=wx.BoxSizer(wx.VERTICAL)
contacts.SetSizer(contactsSizer)
contactsText = wx.TextCtrl(contacts, -1, style=wx.TE_READONLY|wx.TE_MULTILINE)
contactsText.write('''Дмитрий Андреевич Сепселев
Кирилл Викторович Козимянец''')
contactsSizer.Add(contactsText, flag=wx.EXPAND|wx.ALL)
notebook.AddPage(contacts, 'Контакты')
mainFrame.Layout()
mainFrame.Show()
app.MainLoop()