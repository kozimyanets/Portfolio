import sys
import openpyxl
import wx

class choicer(wx.Panel):
    '''Панель с заголовком и раскрывающимся списком для выбора одного из предопределенных вариантов'''
    def __init__(self, heading, variants):
        wx.Panel.__init__(self, panel, -1)
        self.heading=wx.StaticText(self, -1, label=heading, size=wx.Size(200, 30))
        self.heading.SetFont(font)
        self.choice=wx.ComboBox(self, -1, choices=['']+variants, style=wx.CB_READONLY|wx.CB_SORT, size=wx.Size(200, 30))
        self.choice.SetFont(font)
        sizer=wx.BoxSizer(wx.HORIZONTAL)
        self.SetSizer(sizer)
        sizer.Add(self.heading)
        sizer.Add(self.choice)

def update(event):
    '''Функция, которая срабатывает в случае изменения одного из фильтров'''
    filter['department']=departmentChoicer.choice.GetStringSelection()
    filter['object']=objectChoicer.choice.GetStringSelection()
    filter['doc']=docChoicer.choice.GetStringSelection()
    filter['other']=otherChoicer.choice.GetStringSelection()
    listItem=list(base.keys())
    if filter['department']: listItem=[x for x in listItem if base[x][2]==filter['department']]
    if filter['object']: listItem=[x for x in listItem if base[x][3]==filter['object']]
    if filter['doc']: listItem=[x for x in listItem if base[x][4]==filter['doc']]
    if filter['other']:
        a=heading.index(filter['other'])
        listItem=[x for x in listItem if base[x][a-1]==1]
    rpList.DeleteAllItems()
    if listItem==[]: rpText.SetValue('Не найдено предложений, соответствующих всем фильтрам.')
    for x in listItem: rpList.Append((x,))

def changeListItem(event):
    '''Функция, которая срабатывает в случае выбора нового элемента списка предложений'''
    selected=rpList.GetFirstSelected()
    if selected==-1: rpText.Clear()
    else:
        selected=rpList.GetItemText(selected)
        rpText.ChangeValue(base[int(selected)][1])

app=wx.App()
mainWindow=wx.Frame(None, -1, title='Классификатор РП', size=wx.Size(1000, 600))

file='rp.xlsx'#wx.FileSelector(message='Выберете файл', default_path=sys.path[0], parent=mainWindow)
wb=openpyxl.load_workbook(file)
sheet=wb['Лист1']
i=0; j=0
a=1; b=1
while a: i+=1; a=sheet.cell(i, 1).value
while b: j+=1; b=sheet.cell(1, j).value
base={}
for x in range(2, i):
    row=[sheet.cell(x, y).value for y in range(1, j)]
    base[row[0]]=row[1:]
heading=[sheet.cell(1, y).value for y in range(1, j)]


filter = {}
panel=wx.Panel(mainWindow, -1, size=wx.Size(1200, 500))
rpList=wx.ListCtrl(panel, -1, size=wx.Size(100, 500), style=wx.LC_REPORT|wx.LC_SINGLE_SEL)
rpList.AppendColumn('РП')
for x in base.keys(): rpList.Append((x,))
rpText=wx.TextCtrl(panel, -1, size=wx.Size(600, 500), style=wx.TE_READONLY|wx.TE_MULTILINE|wx.TE_CENTRE)
font=rpText.GetFont()
font.SetFractionalPointSize(10)
rpList.SetFont(font)
rpText.SetFont(font)
mainSizer=wx.BoxSizer(wx.HORIZONTAL)
panel.SetSizer(mainSizer)
filterSizer=wx.BoxSizer(wx.VERTICAL)
filterSizer.Add(wx.Size(400, 30))
allDepartments=[base[x][2] for x in base.keys()]
allDepartments=list(set(allDepartments))
departmentChoicer=choicer('Место внедрения: ', allDepartments)
filterSizer.Add(departmentChoicer)
allObjects=[base[x][3] for x in base.keys()]
allObjects=list(set(allObjects))
objectChoicer=choicer('Объект применения: ', allObjects)
filterSizer.Add(objectChoicer)
allDocs=[base[x][4] for x in base.keys()]
allDocs=list(set(allDocs))
docChoicer=choicer('Изменение КД и ТД: ', allDocs)
filterSizer.Add(docChoicer)
otherChoicer=choicer('Другие фильтры: ', heading[6:])
filterSizer.Add(otherChoicer)
panel.Bind(wx.EVT_COMBOBOX, update, departmentChoicer.choice)
panel.Bind(wx.EVT_COMBOBOX, update, objectChoicer.choice)
panel.Bind(wx.EVT_COMBOBOX, update, docChoicer.choice)
panel.Bind(wx.EVT_COMBOBOX, update, otherChoicer.choice)
mainSizer.Add(rpList)
mainSizer.Add(rpText)
mainSizer.Add(wx.Size(50, 500))
mainSizer.Add(filterSizer)
panel.Bind(wx.EVT_LIST_ITEM_SELECTED, changeListItem, rpList)
mainWindow.Layout()
mainWindow.Show()
app.SetTopWindow(mainWindow)
app.MainLoop()